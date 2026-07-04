# -*- coding: utf-8 -*-
"""Fill the official Excel estimate template from estimate-app JSON.

The source template is never modified. A copy is written to outputs/ and then
Excel can export that copy to PDF while preserving the template layout.
"""

from __future__ import annotations

import json
import re
import shutil
import sys
import tempfile
import zipfile
from copy import copy
from datetime import date
from pathlib import Path

from PIL import Image as PillowImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font
from openpyxl.worksheet.properties import PageSetupProperties


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = ROOT / "templates"
OUTPUT_DIR = ROOT / "outputs"
DEFAULT_OUTPUT_WORKBOOK = OUTPUT_DIR / "estimate-template-test.xlsx"
TARGET_SHEET = "裕吏建設"
DEFAULT_CONSTRUCTION_PERMIT = "奈良県知事許可（般-8）第18816号"
SEAL_IMAGE = TEMPLATE_DIR / "takemoto-seal.png"
OFFICIAL_COMPANY_DEFAULTS = {
    "name": "竹本塗装店",
    "representative": "竹本直也",
    "postalCode": "634-0831",
    "address": "奈良県橿原市見瀬町616-2",
    "phone": "0744-35-5162",
    "fax": "0745-51-1854",
    "mobile": "090-2596-1935",
    "email": "takemototosou@gmail.com",
    "registrationNumber": "T2810826383922",
}


def find_template() -> Path:
    candidates = sorted(TEMPLATE_DIR.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"Excel template was not found under {TEMPLATE_DIR}")

    for keyword in ("提出用v2", "提出用テンプレート", "提出用"):
        for candidate in candidates:
            if keyword in candidate.name:
                return candidate
    return candidates[0]


def read_payload(input_path: str | None) -> dict:
    if not input_path:
        return build_test_payload()
    with open(input_path, "r", encoding="utf-8") as file:
        return json.load(file)


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def parse_japanese_date(value: str) -> date | None:
    match = re.search(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", value)
    if not match:
        return None
    return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))


def parse_date_range(value: str | None) -> tuple[date | None, date | None]:
    source = text(value)
    if not source:
        return None, None
    matches = re.findall(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", source)
    if len(matches) >= 2:
        first, second = matches[0], matches[1]
        return (
            date(int(first[0]), int(first[1]), int(first[2])),
            date(int(second[0]), int(second[1]), int(second[2])),
        )
    single = parse_japanese_date(source)
    return single, None


def set_date_parts(ws, prefix_cell: str, value: date | None) -> None:
    col = "".join(ch for ch in prefix_cell if ch.isalpha())
    row = int("".join(ch for ch in prefix_cell if ch.isdigit()))
    if value is None:
        ws[f"{col}{row}"] = None
        ws[f"{col}{row + 1}"] = None
        ws[f"{col}{row + 2}"] = None
        return
    ws[f"{col}{row}"] = value.year
    ws[f"{col}{row + 1}"] = value.month
    ws[f"{col}{row + 2}"] = value.day


def text(value) -> str:
    return "" if value is None else str(value).strip()


def number(value, default=0):
    if value in ("", None):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def yen(value) -> str:
    return f"¥{int(round(number(value))):,}"


def build_company_block(company: dict) -> str:
    return "\n".join(build_company_lines(company))


def official_company_value(company: dict, key: str) -> str:
    value = text(company.get(key))
    if (
        not value
        or "サンプル" in value
        or value in {"000-0000", "06-0000-0000", "info@example.com", "T0000000000000"}
    ):
        return OFFICIAL_COMPANY_DEFAULTS.get(key, "")
    return value


def build_company_lines(company: dict) -> list[str]:
    postal = official_company_value(company, "postalCode")
    address = official_company_value(company, "address")
    lines = [
        official_company_value(company, "name"),
        address,
        f"TEL:{official_company_value(company, 'phone')}　FAX:{official_company_value(company, 'fax')}",
        f"携帯:{official_company_value(company, 'mobile')}",
        f"Mail:{official_company_value(company, 'email')}",
        f"代表：{official_company_value(company, 'representative')}",
    ]
    return [line for line in lines if line]


def build_recipient_block(quote: dict) -> str:
    first_line = f"{text(quote.get('recipientName'))}　{text(quote.get('recipientTitle'))}".strip()
    lines = [first_line]
    if text(quote.get("contactPerson")):
        lines.append(f"ご担当 {text(quote.get('contactPerson'))} 様")
    if text(quote.get("postalCode")):
        lines.append(text(quote.get("postalCode")))
    if text(quote.get("address")):
        lines.append(text(quote.get("address")))
    return "\n".join(line for line in lines if line)


def clear_detail_rows(ws) -> None:
    for row in range(29, 54):
        ws[f"B{row}"] = None
        ws[f"Q{row}"] = None
        ws[f"T{row}"] = None
        ws[f"W{row}"] = None
        ws[f"AG{row}"] = None


def adjust_company_block(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == "Z9:AM13":
            ws.unmerge_cells(str(merged_range))
            break

    ws.merge_cells("Z9:AM9")
    ws.merge_cells("Z10:AM13")

    name_cell = ws["Z9"]
    name_cell.value = None
    name_cell.font = Font(name="メイリオ", size=13, bold=True)
    name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    detail_cell = ws["Z10"]
    detail_cell.value = None
    font = copy(detail_cell.font)
    font.name = "メイリオ"
    font.sz = 9
    detail_cell.font = font
    detail_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def replace_company_card(ws, company: dict, output_path: Path) -> None:
    kept_images = []
    for image in ws._images:
        anchor = image.anchor
        marker = getattr(anchor, "_from", None)
        if marker is None:
            kept_images.append(image)
            continue
        if 24 <= marker.col <= 38 and 7 <= marker.row <= 16:
            continue
        kept_images.append(image)
    ws._images = kept_images

    return


def replace_print_seal(ws) -> None:
    kept_images = []
    for image in ws._images:
        anchor = image.anchor
        marker = getattr(anchor, "_from", None)
        if marker is None:
            kept_images.append(image)
            continue
        # Existing printed shop seal is inside the B:AM print area around AH24.
        if 32 <= marker.col <= 35 and 22 <= marker.row <= 25:
            continue
        kept_images.append(image)
    ws._images = kept_images

    # The official seal image is already transparent. Remove the old template
    # frame so the seal itself can be placed naturally without visible misfit.
    no_border = Border()
    for row in range(24, 28):
        for col in range(34, 40):
            ws.cell(row=row, column=col).border = no_border

    if SEAL_IMAGE.exists():
        seal = ExcelImage(SEAL_IMAGE)
        seal.width = 74
        seal.height = 74
        seal.anchor = "AI24"
        ws.add_image(seal)


def make_embedded_pngs_transparent(workbook_path: Path) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp) / workbook_path.name
        with zipfile.ZipFile(workbook_path, "r") as source, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename.startswith("xl/media/") and item.filename.lower().endswith(".png"):
                    data = make_png_white_transparent(data)
                target.writestr(item, data)
        shutil.move(tmp_path, workbook_path)


def make_png_white_transparent(data: bytes) -> bytes:
    from io import BytesIO

    source = BytesIO(data)
    image = PillowImage.open(source).convert("RGBA")
    pixels = image.load()
    for y in range(image.height):
        for x in range(image.width):
            red, green, blue, alpha = pixels[x, y]
            if alpha and red >= 245 and green >= 245 and blue >= 245:
                pixels[x, y] = (red, green, blue, 0)
    output = BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


def detail_rows_from_quote(quote: dict) -> list[dict]:
    rows = []
    for item in quote.get("items") or []:
        work_item = text(item.get("workItem"))
        description = text(item.get("description"))
        quantity = item.get("quantity")
        unit_price = item.get("unitPrice")
        if not any([work_item, description, text(quantity), text(unit_price)]):
            continue
        rows.append(
            {
                "name": work_item,
                "spec": description,
                "quantity": quantity,
                "unit": text(item.get("unit")) or "式",
                "unit_price": unit_price,
            }
        )

    discount = int(round(number(quote.get("discount"))))
    if discount > 0 and len(rows) < 25:
        rows.append(
            {
                "name": "値引き",
                "spec": "税抜値引き",
                "quantity": 1,
                "unit": "式",
                "unit_price": -discount,
            }
        )
    return rows[:25]


def fill_workbook(payload: dict, output_path: Path) -> Path:
    quote = payload.get("quote") or {}
    company = payload.get("company") or {}
    template = find_template()

    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output_path)

    wb = load_workbook(output_path)
    sheet_name = TARGET_SHEET if TARGET_SHEET in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.print_area = "B2:AM63"
    adjust_company_block(ws)
    replace_company_card(ws, company, output_path)
    replace_print_seal(ws)

    quote_number = text(quote.get("quoteNumber")) or "estimate"
    issue_date = parse_date(quote.get("issueDate"))
    expiry_date = parse_date(quote.get("expiryDate"))
    period_start, period_end = parse_date_range(quote.get("constructionPeriod"))

    ws["B5"] = f"見積書番号：{quote_number}"
    ws["B5"].number_format = "General"
    ws["B8"] = build_recipient_block(quote)
    company_lines = build_company_lines(company)
    ws["Z9"] = company_lines[0] if company_lines else ""
    ws["Z10"] = "\n".join(company_lines[1:])
    ws["K21"] = text(quote.get("subject"))
    ws["K22"] = text(quote.get("siteName"))
    ws["K23"] = text(quote.get("note"))
    if period_start:
        set_date_parts(ws, "AU17", period_start)
        set_date_parts(ws, "BD17", period_end)
    else:
        ws["K24"] = text(quote.get("constructionPeriod"))
        ws["R24"] = None
        ws["S24"] = None
    ws["K25"] = text(quote.get("paymentTerms"))

    ws["AU6"] = "外税"
    ws["AU7"] = 10
    ws["AU9"] = "四捨五入"
    ws["AU15"] = quote_number
    set_date_parts(ws, "AU11", issue_date)
    set_date_parts(ws, "AU21", expiry_date)

    clear_detail_rows(ws)
    for offset, detail in enumerate(detail_rows_from_quote(quote)):
        row = 29 + offset
        ws[f"B{row}"] = detail["name"]
        ws[f"Q{row}"] = number(detail["quantity"], "")
        ws[f"T{row}"] = detail["unit"]
        ws[f"W{row}"] = number(detail["unit_price"], "")
        ws[f"AG{row}"] = detail["spec"]

    discount = int(round(number(quote.get("discount"))))
    rounding = int(round(number(quote.get("roundingAdjustment"))))
    ws["W54"] = "小計"
    ws["W56"] = "合計"
    ws["AA55"] = '=IF(exclusive,ROUND(AA54*taxRate/100,0),"")'
    ws["AA56"] = f"=IF(exclusive,SUM(AA54:AF55),AA54)-{rounding}"
    ws["L17"] = "=AA56"
    ws["L17"].number_format = '"¥"#,##0"－税込";"¥"\\-#,##0"－税込";"¥"0"－税込";@'

    adjustments = []
    if rounding > 0:
        adjustments.append(f"端数調整　-{yen(rounding)}")
    registration_line = "　".join(
        part
        for part in [
            official_company_value(company, "registrationNumber")
            and f"登録番号　{official_company_value(company, 'registrationNumber')}",
            text(company.get("constructionPermit")) or DEFAULT_CONSTRUCTION_PERMIT,
        ]
        if part
    )
    ws["B58"] = "\n".join(line for line in [registration_line, "　　".join(adjustments)] if line)

    wb.active = wb.sheetnames.index(sheet_name)
    wb.save(output_path)
    make_embedded_pngs_transparent(output_path)
    print(json.dumps({"output": str(output_path), "sheet": sheet_name}, ensure_ascii=False))
    return output_path


def build_test_payload() -> dict:
    return {
        "quote": {
            "quoteNumber": "TKM-2026-001",
            "recipientName": "テスト建設株式会社",
            "recipientTitle": "御中",
            "subject": "テスト邸 外壁・屋根塗装工事",
            "siteName": "奈良県橿原市テスト町1-2-3",
            "constructionPeriod": "2026年7月1日 ～ 2026年7月15日",
            "issueDate": "2026-07-03",
            "expiryDate": "2026-08-02",
            "paymentTerms": "工事完了後、請求書発行月の翌月末までにお支払いください。",
            "items": [
                {"workItem": "仮設足場", "description": "足場・シート含む", "quantity": 1, "unit": "式", "unitPrice": 180000},
                {"workItem": "高圧洗浄", "description": "外壁・屋根", "quantity": 200, "unit": "㎡", "unitPrice": 250},
                {"workItem": "外壁下地補修", "description": "クラック補修", "quantity": 1, "unit": "式", "unitPrice": 85000},
                {"workItem": "外壁下塗り", "description": "下塗り材", "quantity": 200, "unit": "㎡", "unitPrice": 700},
                {"workItem": "外壁中塗り・上塗り", "description": "シリコン塗料", "quantity": 200, "unit": "㎡", "unitPrice": 2000},
            ],
        },
        "company": {
            "name": "竹本塗装店",
            "representative": "竹本直也",
            "postalCode": "634-0831",
            "address": "奈良県橿原市見瀬町616-2",
            "phone": "0744-35-5162",
            "registrationNumber": "T2810826383922",
        },
    }


if __name__ == "__main__":
    input_json = sys.argv[1] if len(sys.argv) > 1 else None
    output_xlsx = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT_WORKBOOK
    fill_workbook(read_payload(input_json), output_xlsx)
